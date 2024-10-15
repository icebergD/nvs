import decimal

import openpyxl

def theme2rgb(theme_color, tint_color):
    theme_colors = [
        ['ffffff', '000000', 'eeece1', '1f497d', '4f81bd', 'c0504d', '9bbb59', '8064a2', '4bacc6', 'f79646'],
        ['f2f2f2', '7f7f7f', 'ddd9c3', 'c6d9f0', 'dbe5f1', 'f2dcdb', 'ebf1dd', 'e5e0ec', 'dbeef3', 'fdeada'],
        ['d8d8d8', '595959', 'c4bd97', '8db3e2', 'b8cce4', 'e5b9b7', 'd7e3bc', 'ccc1d9', 'b7dde8', 'fbd5b5'],
        ['bfbfbf', '3f3f3f', '938953', '548dd4', '95b3d7', 'd99694', 'c3d69b', 'b2a2c7', '92cddc', 'fac08f'],
        ['a5a5a5', '262626', '494429', '17365d', '366092', '953734', '76923c', '5f497a', '31859b', 'e36c09'],
        ['7f7f7f', '0c0c0c', '1d1b10', '0f243e', '244061', '632423', '4f6128', '3f3151', '205867', '974806']
    ]
    if str(theme_color)[0:6] == 'Values':
        return None
    tint_color_index = 0
    tint_color = round(tint_color*(100))
    if tint_color == 0:
        tint_color_index = 0
    elif tint_color == 80:
        tint_color_index = 1
    elif tint_color == 60:
        tint_color_index = 2
    elif tint_color == 40:
        tint_color_index = 3
    elif tint_color == -25:
        tint_color_index = 4
    elif tint_color == -50:
        tint_color_index = 5
    return theme_colors[tint_color_index][theme_color]


def extract_cells_from_merged_ranges(merged_ranges):
    cells = []
    num_ranges = len(merged_ranges)
    for i, merged_range in enumerate(merged_ranges):
        start_row, start_col = merged_range.min_row, merged_range.min_col
        end_row, end_col = merged_range.max_row, merged_range.max_col

        if start_row == end_row and start_col == end_col:
            # Range contains only one cell, so skip it
            continue

        # Add all cells in the range, except for the starting cell reference
        for row in range(start_row, end_row+1):
            for col in range(start_col, end_col+1):
                if row == start_row and col == start_col:
                    # Skip the starting cell reference
                    continue
                cells.append((col, row))

    return cells


def cell_get_params(sheet):
    merged_cells = sheet.merged_cells.ranges
    exclude_cells = extract_cells_from_merged_ranges(merged_cells)
    sheet_arr = []
    char_line = []
    num_line = []
    # Get the row count and column count
    row_count = sheet.max_row
    column_count = sheet.max_column

    freeze_panes = sheet.freeze_panes
    freeze_row = 0
    if freeze_panes != None:
        freeze_panes = freeze_panes.split(':')
        freeze_row = int(freeze_panes[0][1:])
    # print(freeze_row)

    for el in range(1, column_count+1):
        col_value = openpyxl.utils.get_column_letter(el)
        # Get the width of column
        col_width = sheet.column_dimensions[col_value].width
        char_line.append({'col_value':col_value, 'col_index':el, 'col_width':col_width})

    for el in range(1, row_count+1):
        # Get the width of column
        row_height = sheet.row_dimensions[el].height
        num_line.append({'row_value': el, 'row_index': el, 'row_height':row_height, 'exists': True})

    for i , n_l in zip(sheet, num_line):
        row_arr =[]
        row_arr.append(n_l)
        for cell in i:
            row = cell.row
            col = cell.column
            value = cell.value
            if value == None:
                value = ''
            elif str(value)[0] == '=':
                value = ''
            coordinate = cell.coordinate

            #background
            fill = cell.fill
            theme_color = fill.start_color.theme
            tint_color = fill.start_color.tint
            bg_color1 = theme2rgb(theme_color, tint_color)
            bg_color2 = str(fill.start_color.rgb)
            background_color = 'ffffff'
            if bg_color1 != None:
                background_color = bg_color1
            elif bg_color2 != '00000000':
                background_color = bg_color2[2:8]
            background_color = '#' + background_color


            font = cell.font
            font_name = font.name
            font_bold = font.b
            font_italic = font.i
            font_color = '#000000' #font.color
            font_size = font.size

            #borders
            border_top_val = '1'
            border_top_color = '#c4c4c4'
            border_top = cell.border.top.style
            if border_top == 'medium':
                border_top_val = '2'
                border_top_color = 'black'
            elif border_top == 'thin':
                border_top_val = '1'
                border_top_color = 'black'
            border_right_val = '1'
            border_right_color = '#c4c4c4'
            border_right = cell.border.right.style
            if border_right == 'medium':
                border_right_val = '2'
                border_right_color = 'black'
            elif border_right == 'thin':
                border_right_val = '1'
                border_right_color = 'black'
            border_bottom_val = '1'
            border_bottom_color = '#c4c4c4'
            border_bottom = cell.border.bottom.style
            if border_bottom == 'medium':
                border_bottom_val = '2'
                border_bottom_color = 'black'
            elif border_bottom == 'thin':
                border_bottom_val = '1'
                border_bottom_color = 'black'
            border_left_val = '1'
            border_left_color = '#c4c4c4'
            border_left = cell.border.left.style
            if border_left == 'medium':
                border_left_val = '2'
                border_left_color = 'black'
            elif border_left == 'thin':
                border_left_val = '1'
                border_left_color = 'black'

            alignment = cell.alignment
            horizontal_align = alignment.horizontal
            vertical_align = alignment.vertical
            wrap_text = alignment.wrapText

            number_format = cell.number_format

            protection = cell.protection

            # Check if the cell is part of a merged cell range
            merged_cells = sheet.merged_cells.ranges
            # print(merged_cells)
            visable = True
            colspan = 1
            rowspan = 1
            for merged_cell in merged_cells:
                # Get the first cell in the merged cell range
                first_cell = merged_cell.min_col, merged_cell.min_row
                # Get the last cell in the merged cell range
                last_cell = merged_cell.max_col, merged_cell.max_row
                # Check if the current cell is the first cell in the range
                if (col, row) == first_cell:
                    colspan = last_cell[0] - first_cell[0] + 1
                    rowspan = last_cell[1] - first_cell[1] + 1
            #убираем лишние ячейки
            for exclude_cell in exclude_cells:
                if (col, row) == exclude_cell:
                    visable = False

            frozen = False
            if row < freeze_row:
                frozen = True
            cell_params = {
                'value': value,
                'row': row,
                'col': col,
                'coordinate': coordinate,
                'colspan': colspan,
                'rowspan': rowspan,
                'visable': visable,
                'font_name': font_name,
                'font_bold': font_bold,
                'font_italic': font_italic,
                'font_color': font_color,
                'font_size': font_size,
                'background_color': background_color,

                'border_top': border_top_val,
                'border_right': border_right_val,
                'border_bottom': border_bottom_val,
                'border_left': border_left_val,
                'border_top_color': border_top_color,
                'border_right_color': border_right_color,
                'border_bottom_color': border_bottom_color,
                'border_left_color': border_left_color,

                'horizontal_align': horizontal_align,
                'vertical_align': vertical_align,
                'wrap_text': wrap_text,

                'frozen': frozen,

                'number_format': number_format,
            }
            row_arr.append(cell_params)
            # Print the parameters
            # print("Row: ", row)
            # print("Column: ", col)
            # print("Value: ", value)
            #

            # print("Number Format: ", number_format)
            # print("Protection: ", protection)
            # print('_________________________')
        sheet_arr.append(row_arr)
    return sheet_arr, char_line, num_line


# def format_excel_style(value, number_format):
#     """
#     Formats a value using an Excel-style number format.
#
#     Parameters:
#     value (float): The value to be formatted.
#     number_format (str): The Excel-style number format to be used for formatting.
#
#     Returns:
#     str: The formatted string.
#     """
#     import openpyxl
#
#     wb = openpyxl.Workbook()
#     ws = wb.active
#
#     cell = ws.cell(row=1, column=1)
#     cell.number_format = number_format
#     cell.value = value
#
#     formatted_value = cell.value
#
#     return formatted_value
#
#
#
# def format_value(value, number_format):
#     if number_format=='General':
#         return value
#     elif '%' in number_format:
#         return value
#     elif '#' in number_format:
#         return value
#     return value
#
#
# import re
# from decimal import Decimal
#
#
# def format_excel_number(value, format_string):
#     # convert value to a Decimal object
#     try:
#         value = Decimal(str(value))
#     except decimal.InvalidOperation:
#         return value
#     # parse the format string
#     format_parts = re.findall(r'[^\d.#]+|[\d.#]+', format_string)
#
#     # set defaults for format options
#     decimal_places = 0
#     thousands_separator = ''
#     decimal_separator = '.'
#     positive_prefix = ''
#     positive_suffix = ''
#     negative_prefix = '-'
#     negative_suffix = ''
#
#     # parse the format options
#     for part in format_parts:
#         if part.startswith('#'):
#             decimal_places = len(part) - 1
#         elif part.startswith('0'):
#             decimal_places = len(part) - 1
#         elif part == ',':
#             thousands_separator = ','
#         elif part == '.':
#             decimal_separator = '.'
#         elif part.startswith('$'):
#             positive_prefix = '$'
#         elif part.endswith('$'):
#             positive_suffix = '$'
#         elif part.startswith('('):
#             negative_prefix = '('
#             negative_suffix = ')'
#         elif part.endswith(')'):
#             negative_suffix = ')'
#
#     # format the value
#     if value >= 0:
#         prefix = positive_prefix
#         suffix = positive_suffix
#     else:
#         prefix = negative_prefix
#         suffix = negative_suffix
#         value = abs(value)
#
#     formatted_value = f"{value:,.{decimal_places}f}".replace(',', thousands_separator).replace('.', decimal_separator)
#
#     return f"{prefix}{formatted_value}{suffix}"
#
# import re
#
# def excel_format(value, format_string):
#     if format_string != 'General':
#         # Replace any Excel-style format codes with their equivalent Python format codes
#         format_string = re.sub(r'(?<!\\)#', '{}', format_string)
#         format_string = re.sub(r'(?<!\\)0', '{:0}', format_string)
#         format_string = re.sub(r'(?<!\\)\.', '{:.}', format_string)
#         format_string = re.sub(r'(?<!\\)%', '{:.0%}', format_string)
#         format_string = re.sub(r'(?<!\\)\?', '{}', format_string)
#         format_string = re.sub(r'(?<!\\)g', '{:f}', format_string)
#
#         # Escape any remaining backslashes
#         format_string = format_string.replace('\\', '\\\\')
#
#         # Format the value using the modified format string
#         # return format_string.format(value)
#         return format_string
#     else:
#         return value

def is_in_array(cell, arr):#, user, excel_table_id):
    c_row = cell['row']
    c_col = cell['col']
    c_sheet = 'tab'+str(cell['sheet'])
    for i in arr:
        if i['sheet'] == c_sheet and i['row'] == c_row and i['col'] == c_col:
            return True
    return False

def convert_table_to_client(arr, cells):
    # print(cells.values())
    for i in arr:
        for j in i[1:]:
            for k in cells:
                if j['row'] == k.row and j['col'] == k.col:
                    j['input'] = True
    # print(arr)
    return arr

def convert_table_to_view(arr, cells):
    # print(cells.values())
    for i in arr:
        for j in i[1:]:
            for k in cells:
                if j['row'] == k.row and j['col'] == k.col:
                    j['value'] = k.value
    return arr

def cell_get_params(sheet):
    merged_cells = sheet.merged_cells.ranges
    exclude_cells = extract_cells_from_merged_ranges(merged_cells)
    sheet_arr = []
    char_line = []
    num_line = []
    # Get the row count and column count
    row_count = sheet.max_row
    column_count = sheet.max_column

    freeze_panes = sheet.freeze_panes
    freeze_row = 0
    if freeze_panes != None:
        freeze_panes = freeze_panes.split(':')
        freeze_row = int(freeze_panes[0][1:])
    # print(freeze_row)

    for el in range(1, column_count+1):
        col_value = openpyxl.utils.get_column_letter(el)
        # Get the width of column
        col_width = sheet.column_dimensions[col_value].width*2
        char_line.append({'col_value': col_value, 'col_index':el, 'col_width':col_width})

    for el in range(1, row_count+1):
        # Get the width of column
        row_height = sheet.row_dimensions[el].height
        num_line.append({'row_value': el, 'row_index': el, 'row_height':row_height, 'exists': True})

    for i, n_l in zip(sheet, num_line):
        row_arr = []
        row_arr.append(n_l)
        for cell in i:
            row = cell.row
            col = cell.column
            value = cell.value
            if value == None:
                value = ''
            elif str(value)[0] == '=':
                value = ''
            coordinate = cell.coordinate

            #background
            fill = cell.fill
            theme_color = fill.start_color.theme
            tint_color = fill.start_color.tint
            bg_color1 = theme2rgb(theme_color, tint_color)
            bg_color2 = str(fill.start_color.rgb)
            background_color = 'ffffff'
            if bg_color1 != None:
                background_color = bg_color1
            elif bg_color2 != '00000000':
                background_color = bg_color2[2:8]
            background_color = '#' + background_color


            font = cell.font
            font_name = font.name
            font_bold = font.b
            font_italic = font.i
            font_color = '#000000' #font.color
            font_size = font.size

            #borders
            border_top_val = '1'
            border_top_color = '#c4c4c4'
            border_top = cell.border.top.style
            if border_top == 'medium':
                border_top_val = '2'
                border_top_color = 'black'
            elif border_top == 'thin':
                border_top_val = '1'
                border_top_color = 'black'
            border_right_val = '1'
            border_right_color = '#c4c4c4'
            border_right = cell.border.right.style
            if border_right == 'medium':
                border_right_val = '2'
                border_right_color = 'black'
            elif border_right == 'thin':
                border_right_val = '1'
                border_right_color = 'black'
            border_bottom_val = '1'
            border_bottom_color = '#c4c4c4'
            border_bottom = cell.border.bottom.style
            if border_bottom == 'medium':
                border_bottom_val = '2'
                border_bottom_color = 'black'
            elif border_bottom == 'thin':
                border_bottom_val = '1'
                border_bottom_color = 'black'
            border_left_val = '1'
            border_left_color = '#c4c4c4'
            border_left = cell.border.left.style
            if border_left == 'medium':
                border_left_val = '2'
                border_left_color = 'black'
            elif border_left == 'thin':
                border_left_val = '1'
                border_left_color = 'black'

            alignment = cell.alignment
            horizontal_align = alignment.horizontal
            vertical_align = alignment.vertical
            wrap_text = alignment.wrapText

            number_format = cell.number_format

            protection = cell.protection

            # Check if the cell is part of a merged cell range
            merged_cells = sheet.merged_cells.ranges
            # print(merged_cells)
            visable = True
            colspan = 1
            rowspan = 1
            for merged_cell in merged_cells:
                # Get the first cell in the merged cell range
                first_cell = merged_cell.min_col, merged_cell.min_row
                # Get the last cell in the merged cell range
                last_cell = merged_cell.max_col, merged_cell.max_row
                # Check if the current cell is the first cell in the range
                if (col, row) == first_cell:
                    colspan = last_cell[0] - first_cell[0] + 1
                    rowspan = last_cell[1] - first_cell[1] + 1
            #убираем лишние ячейки
            for exclude_cell in exclude_cells:
                if (col, row) == exclude_cell:
                    visable = False

            frozen = False
            if row < freeze_row:
                frozen = True
            cell_params = {
                'value': value,
                'row': row,
                'col': col,
                'coordinate': coordinate,
                'colspan': colspan,
                'rowspan': rowspan,
                'visable': visable,
                'font_name': font_name,
                'font_bold': font_bold,
                'font_italic': font_italic,
                'font_color': font_color,
                'font_size': font_size,
                'background_color': background_color,

                'border_top': border_top_val,
                'border_right': border_right_val,
                'border_bottom': border_bottom_val,
                'border_left': border_left_val,
                'border_top_color': border_top_color,
                'border_right_color': border_right_color,
                'border_bottom_color': border_bottom_color,
                'border_left_color': border_left_color,

                'horizontal_align': horizontal_align,
                'vertical_align': vertical_align,
                'wrap_text': wrap_text,

                'frozen': frozen,

                'number_format': number_format,
            }
            row_arr.append(cell_params)
            # Print the parameters
            # print("Row: ", row)
            # print("Column: ", col)
            # print("Value: ", value)
            #

            # print("Number Format: ", number_format)
            # print("Protection: ", protection)
            # print('_________________________')
        sheet_arr.append(row_arr)
    return sheet_arr, char_line, num_line

def value_convert(value, type):
    try:
        if type == 'General':
            try:
                value = int(value)
            except:
                pass
        if '.' in type:
            value = float(value)
        else:
            try:
                value = int(value)
            except:
                pass
        if '%' in type:
            value = value/100
    except:
        print('error value_convert')
    return value
