import json
import os.path

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.views import View
import io
from .models import ExcelTable, Cell, ExcelTableFinish, CellInput
import openpyxl
import urllib.parse
from .utilitys import cell_get_params, is_in_array, convert_table_to_client, convert_table_to_view, value_convert
from users.mixins import UserStatus
from django.contrib.auth import get_user_model
from datetime import datetime, timezone

class ExcelTableList(UserStatus, View):
    def get(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':#creator or user creator
            excel_table = ExcelTable.objects.filter(user=request.user).order_by('-created')
            table_data = []

            for i in excel_table:
                finished = ExcelTableFinish.objects.filter(table__id=i.id)
                finished_count = len(finished)
                finished_selected = ExcelTableFinish.objects.filter(table__id=i.id, finished=None)
                finished_selected_count = len(finished_selected)
                if finished_selected_count == 0 and finished_count != 0:
                    color = 'bg-success'
                elif finished_count == 0:
                    color = 'bg-danger'
                else:
                    color = 'bg-danger'
                checked_str = ''
                if i.visable==True:
                    checked_str = 'checked'
                table_el = {
                    'id': i.id,
                    'file': os.path.basename(str(i.file)),
                    'description': i.description,
                    'status': f'{finished_count - finished_selected_count}/{finished_count}',
                    'status_color': color,
                    'checked': checked_str,
                    'created': i.created
                }
                table_data.append(table_el)
            return render(request, "excel_table/excelTableList.html", {'table': table_data, 'navbar': 'excel-table-list', 'status': self.m_user, 'user_name': self.user_name})
        else:
            return redirect('login')


class UploadExcel(UserStatus, View):
    def post(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':#creator or user creator
            description = request.POST.get('description')
            excel_file = request.FILES['excel_file']
            print(excel_file)
            excel = ExcelTable.objects.create(
                file=excel_file,
                user=request.user,
                description=description
            )
            return redirect('excel-table-list')
            # return redirect('view_excel', pk=excel.pk)
        else:
            return redirect('login')


class ConfigureExcel(UploadExcel, View):
    def get(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':#creator or user creator
            pk = kwargs.get('pk')
            excel = ExcelTable.objects.get(pk=pk)
            workbook = openpyxl.load_workbook(excel.file)

            active_sheet = workbook.active.title
            sheet_names = workbook.sheetnames
            sheets = []
            active_sheet_id = ''
            for index, sheet_name in enumerate(sheet_names):
                sheet = workbook[sheet_name]
                sheet_arr, char_line, num_line = cell_get_params(sheet)
                el = {
                    'id': 'tab'+str(index),
                    'sheet': sheet_arr,
                    'char_line': char_line,
                    'num_line': num_line,
                    'sheet_name': sheet_name
                }
                sheets.append(el)
                if active_sheet == sheet_name:
                    active_sheet_id = 'tab'+str(index)

            User = get_user_model()
            users = User.objects.all().values('id', 'first_name');

            cells_obj = Cell.objects.filter(excel_table__id=pk).values()
            users_selected_cells = []
            selected_users = []
            for i in cells_obj:
                if i['cell_user_id'] not in selected_users:
                    selected_users.append(i['cell_user_id'])

            for i in selected_users:
                cells = Cell.objects.filter(excel_table__id=pk, cell_user_id=i).values()
                temp_arr = []
                for j in cells:
                    temp_arr.append({'sheet': 'tab'+str(j['sheet']), 'row': j['row'], 'col': j['col']})
                users_selected_cells.append(temp_arr)


            for i, value in enumerate(selected_users):
                for j in users:
                    if value == j['id']:
                        selected_users[i] = j['first_name']

            # print(users_selected_cells)
            # print(selected_users)
            return render(request, 'excel_table/configure_excel.html', {'sheets': sheets, 'active_sheet': active_sheet, 'active_sheet_id': active_sheet_id, 'status': self.m_user, 'users': users, 'users_selected_cells': users_selected_cells, 'selected_users': selected_users})
        else:
            return redirect('login')

    def post(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':#creator or user creator
            #excel_table = list(ExcelTable.objects.filter(user=request.user).order_by('-created').values())
            User = get_user_model()

            users_arr = json.loads(request.POST.get('users'))
            user_cell_arr = json.loads(request.POST.get('users_cells'))

            excel_table_id = kwargs.get('pk')
            excel_table = ExcelTable.objects.get(pk=excel_table_id)
            for user, cells_new in zip(users_arr, user_cell_arr):
                user_obj = User.objects.get(first_name=user)
                cells_exist = Cell.objects.filter(excel_table=excel_table, cell_user=user_obj).values()

                # поиск и удаление ячеек, которые не отмечены
                for cell in cells_exist:
                    is_exist_cell = is_in_array(cell, cells_new)
                    # print(is_exist_cell)
                    if not is_exist_cell:
                        Cell.objects.filter(id=cell['id']).delete()
                        # print(cell)




                # удаление заполненных ячеек
                CellInput.objects.filter(excel_table=excel_table, cell_user=user_obj).delete()

                for cell in cells_new:
                    c_row = cell['row']
                    c_col = cell['col']
                    c_sheet = int(cell['sheet'][3:])
                    is_exist_cell = list(Cell.objects.filter(excel_table=excel_table, cell_user=user_obj, row=c_row, col=c_col, sheet=c_sheet).values())
                    if len(is_exist_cell) == 0:
                        cell_obj = Cell()
                        cell_obj.cell_user = user_obj
                        cell_obj.excel_table = excel_table

                        cell_obj.row = c_row
                        cell_obj.col = c_col
                        cell_obj.sheet = c_sheet
                        cell_obj.save()

                try:
                    excel_table_finfish_obj = ExcelTableFinish.objects.get(table=excel_table, user=user_obj)
                    excel_table_finfish_obj.finished = None
                    excel_table_finfish_obj.save()
                except ExcelTableFinish.DoesNotExist:# если не существует, создать
                    excel_table_finfish = ExcelTableFinish()
                    excel_table_finfish.table = excel_table
                    excel_table_finfish.user = user_obj
                    excel_table_finfish.save()


            # если нету пользователя то удалить все ячейки этого пользователя
            all_cells = Cell.objects.filter(excel_table=excel_table).all()
            for i in all_cells:
                if i.cell_user.first_name not in users_arr:
                    i.delete()

            all_finish = ExcelTableFinish.objects.filter(table=excel_table).all()
            for i in all_finish:
                if i.user.first_name not in users_arr:
                    i.delete()

            return JsonResponse({'url': '/excel/excel-table-list'})
        else:
            return redirect('login')


# def upload_excel(request):
#     if request.method == 'POST':
#         excel_file = request.FILES['excel_file']
#         excel = ExcelTable.objects.create(file=excel_file)
#         return redirect('view_excel', pk=excel.pk)
#     return render(request, 'excel_table/upload_excel.html')

#не нужно
def view_excel(request, pk):
    excel = ExcelTable.objects.get(pk=pk)
    workbook = openpyxl.load_workbook(excel.file)
    sheets = workbook.sheetnames
    return render(request, 'excel_table/view_excel.html', {'sheets': sheets, 'excel': excel})

#не нужно
def view_sheet(request, pk, sheet_name):
    excel = ExcelTable.objects.get(pk=pk)
    workbook = openpyxl.load_workbook(excel.file)
    sheet = workbook[sheet_name]
    sheet_arr, char_line, num_line = cell_get_params(sheet)
    return render(request, 'excel_table/view_sheet.html', {'sheet': sheet_arr, 'char_line': char_line, 'num_line': num_line})


class ExcelTableToggleVisable(UserStatus, View):
    def get(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':#creator or user creator
            table_id = request.GET.get('val')
            checked = request.GET.get('checked')
            if checked == 'true':
                checked = True
            else:
                checked = False
            ExcelTable.objects.filter(id=table_id).update(visable=checked)

            return JsonResponse({'done':'done'})
        else:
            return redirect('login')

class DeleteExcelTable(UserStatus, View):
    def post(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':#creator or user creator
            id = request.POST.get("val")
            ExcelTable.objects.filter(id=id, user=request.user).delete()
            # table = Table.objects.order_by('-created')
            return redirect('excel-table-list')
        else:
            return redirect('login')


class ClientExcelTableList(UserStatus, View):
    def get(self, request, *args, **kwargs):
        if self.m_user == 't' or self.m_user == 'c' or self.m_user == 'u':#user creator(tanos) or creator or user

            finished = ExcelTableFinish.objects.filter(user=request.user).order_by('finished')
            table_data = []
            for i in finished:

                table = ExcelTable.objects.get(id=i.table.id)
                if not table.visable:
                    continue

                creator = table.user.first_name
                if i.finished != None:
                    color = 'bg-success'
                    status = '+'
                else:
                    color = 'bg-danger'
                    status = '-'
                permision_el = {
                    'id': table.id,
                    'name': os.path.basename(str(table.file)),
                    'description': table.description,
                    'creator': creator,
                    'status': status,
                    'status_color': color,
                    'created': table.created.strftime("%d.%m.%Y \n %H:%M")
                }
                table_data.append(permision_el)
                # print(table_data)
            return render(request, "excel_table/clientExcelTableList.html", {'table': table_data, 'navbar':'client-excel-table-list', 'status': self.m_user, 'user_name': self.user_name})
        else:
            return redirect('login')


class ClientExcelFill(UserStatus, View):
    def get(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't' or self.m_user == 'u':  # creator or user creator or user
            pk = kwargs.get('pk')
            excel = ExcelTable.objects.get(pk=pk)
            workbook = openpyxl.load_workbook(excel.file)

            active_sheet = workbook.active.title
            sheet_names = workbook.sheetnames
            sheets = []
            active_sheet_id = ''
            for index, sheet_name in enumerate(sheet_names):
                sheet = workbook[sheet_name]
                sheet_arr, char_line, num_line = cell_get_params(sheet)
                cells = Cell.objects.filter(excel_table_id=pk, cell_user=request.user, sheet=index)

                sheet_arr = convert_table_to_client(sheet_arr, cells)
                el = {
                    'id': 'tab' + str(index),
                    'sheet': sheet_arr,
                    'char_line': char_line,
                    'num_line': num_line,
                    'sheet_name': sheet_name
                }
                sheets.append(el)
                if active_sheet == sheet_name:
                    active_sheet_id = 'tab' + str(index)


            return render(request, 'excel_table/clientExcelFill.html',
                          {'sheets': sheets, 'active_sheet': active_sheet, 'active_sheet_id': active_sheet_id,
                           'status': self.m_user, 'table':pk})
        else:
            return redirect('login')

    def post(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't' or self.m_user == 'u':  # creator or user creator of user
            values = json.loads(request.POST.get('input_values'))
            # print(values)
            excel_table_id = kwargs.get('pk')
            excel_table = ExcelTable.objects.get(pk=excel_table_id)
            # print(values)
            for i in values:
                cell_input = CellInput()
                cell_input.value = i['value']
                cell_input.excel_table = excel_table
                cell_input.cell_user = request.user
                cell_input.row = i['row']
                cell_input.col = i['col']
                cell_input.sheet = int(i['sheet'][3:])
                cell_input.save()


            # обозначаем что пользователь закончил
            finished = ExcelTableFinish.objects.get(table=excel_table, user=request.user)
            finished.finished = datetime.now(timezone.utc)
            finished.save()

            return JsonResponse({'url': '/excel/client-excel-table-list'})
        else:
            return redirect('login')

class ExcelTableView(UserStatus, View):
    def get(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':  # creator or user creator
            pk = kwargs.get('pk')
            excel = ExcelTable.objects.get(pk=pk)
            workbook = openpyxl.load_workbook(excel.file)

            active_sheet = workbook.active.title
            sheet_names = workbook.sheetnames
            sheets = []
            active_sheet_id = ''
            for index, sheet_name in enumerate(sheet_names):
                sheet = workbook[sheet_name]
                sheet_arr, char_line, num_line = cell_get_params(sheet)
                cells_inputed = CellInput.objects.filter(excel_table_id=pk, sheet=index)

                sheet_arr = convert_table_to_view(sheet_arr, cells_inputed)
                el = {
                    'id': 'tab' + str(index),
                    'sheet': sheet_arr,
                    'char_line': char_line,
                    'num_line': num_line,
                    'sheet_name': sheet_name
                }
                sheets.append(el)
                if active_sheet == sheet_name:
                    active_sheet_id = 'tab' + str(index)

            return render(request, 'excel_table/excelTableView.html',
                          {'sheets': sheets, 'active_sheet': active_sheet, 'active_sheet_id': active_sheet_id,
                           'status': self.m_user})
        else:
            return redirect('login')

    def post(self, request, *args, **kwargs):
        if self.m_user == 'c' or self.m_user == 't':  # creator or user creator
            excel_table_id = kwargs.get('pk')
            excel_table = ExcelTable.objects.get(pk=excel_table_id)

            cell_input = CellInput.objects.filter(excel_table=excel_table)
            sheet_data = {}

            for d in cell_input:
                sheet = d.sheet
                if sheet not in sheet_data:
                    sheet_data[sheet] = []
                sheet_data[sheet].append({'row': d.row, 'col': d.col, 'value': d.value, 'sheet': d.sheet})

            wb = openpyxl.load_workbook(excel_table.file.name)
            for sheet in sheet_data.items():
                ws = wb.worksheets[sheet[0]]
                for item in sheet[1]:
                    row = item['row']
                    col = item['col']
                    value = item['value']
                    cell = ws.cell(row=row, column=col)
                    cell.value = value_convert(value, cell.number_format)

            output = io.BytesIO()
            wb.save(output)

            filename = os.path.basename(excel_table.file.name)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.getvalue()
            )
            response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(urllib.parse.quote(filename))

            return response
            # return redirect('excel-table-list')
        else:
            return redirect('login')